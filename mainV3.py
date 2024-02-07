import sys
from tkinter import filedialog, messagebox
import re
import fitz
from openpyxl.styles import Font
from datetime import datetime
import os
from openpyxl import load_workbook,  Workbook


def extrair_palavra_apos(texto, palavra_chave):
    posicao_palavra_chave = texto.find(palavra_chave)
    if posicao_palavra_chave != -1:
        inicio_palavra = posicao_palavra_chave + \
            len(palavra_chave) + 1  # Adiciona 1 para pular o espaço
        fim_palavra = texto.find(' ', inicio_palavra)
        if fim_palavra == -1:
            palavra_apos = texto[inicio_palavra:]
        else:
            palavra_apos = texto[inicio_palavra:fim_palavra]
        return palavra_apos.strip()  # Adiciona strip() para remover espaços extras
    else:
        return None


def extrair_palavra_apos_ss(texto, palavra_chave):
    posicao_palavra_chave = texto.find(palavra_chave)

    if posicao_palavra_chave != -1:
        inicio_palavra = posicao_palavra_chave + len(palavra_chave)
        fim_palavra = texto.find(' ', inicio_palavra)

        if fim_palavra == -1:
            palavra_apos = texto[inicio_palavra:]
        else:
            palavra_apos = texto[inicio_palavra:fim_palavra]

        return palavra_apos.strip()
    else:
        return None


def extrair_numeros_apos_ss(texto, palavra_chave):
    palavra_apos = extrair_palavra_apos_ss(texto, palavra_chave)
    if palavra_apos:
        numeros = ''.join(c for c in palavra_apos if c.isdigit() or c == '.')
        return numeros if numeros else None
    else:
        return None


def extrair_numeros_apos(texto, palavra_chave):
    palavra_apos = extrair_palavra_apos(texto, palavra_chave)
    if palavra_apos:
        numeros = ''.join(c for c in palavra_apos if c.isdigit() or c == '.')
        return numeros if numeros else None
    else:
        return None


def extrair_numeros_especificos_apos(texto, palavra_chave):
    palavra_apos = extrair_palavra_apos(texto, palavra_chave)
    if palavra_apos:
        numeros = ''.join(c for c in palavra_apos if c.isdigit())
        return numeros[:5] if numeros else None
    else:
        return None


def extrair_proximos_numeros(texto, palavra_chave):
    palavra_apos = extrair_palavra_apos(texto, palavra_chave)
    if palavra_apos:
        numeros = ''.join(c for c in palavra_apos if c.isdigit())
        return numeros[5:11] if len(numeros) >= 11 else None
    else:
        return None


def extrair_outro_dado(texto, palavra_chave):
    palavra_apos = extrair_palavra_apos(texto, palavra_chave)
    return palavra_apos if palavra_apos else None


def extrair_texto_pdf(pdf_path):
    pdf_doc = fitz.open(pdf_path)
    texto_ocr = ""

    for pagina_num in range(pdf_doc.page_count):
        pagina = pdf_doc[pagina_num]
        texto_ocr += pagina.get_text()

    pdf_doc.close()
    return texto_ocr


try:
    pasta_selecionada = filedialog.askdirectory(
        title="SELECIONE A PASTA QUE CONTENHA OS PDFs")

    if not pasta_selecionada:
        raise FileNotFoundError("NENHUMA PASTA SELECIONADA.")

    if not os.path.exists(pasta_selecionada):
        raise FileNotFoundError(
            f"A PASTA selecionada '{pasta_selecionada}' não existe.")

    nomes_arquivos = os.listdir(pasta_selecionada)

    if not nomes_arquivos:
        raise FileNotFoundError(
            f"A PASTA selecionada ('{pasta_selecionada}') está VAZIA.")

    if not all(nome.lower().endswith(".pdf") for nome in nomes_arquivos):
        raise FileNotFoundError(
            f"A PASTA selecionada ('{pasta_selecionada}') deve conter apenas arquivos PDFs.")

    # Restante do seu código aqui...
    for nome_arquivo_pdf in nomes_arquivos:
        caminho_completo_pdf = os.path.join(
            pasta_selecionada, nome_arquivo_pdf)

        try:
            pdf_doc = fitz.open(caminho_completo_pdf)
        except fitz.FilePDFError:
            messagebox.showerror(
                "ERRO", f"O arquivo ('{nome_arquivo_pdf}') na pasta ('{pasta_selecionada}') não é um PDF válido. Ignorando.")
            continue

        texto = extrair_texto_pdf(caminho_completo_pdf)

        palavra_chave_obrigatoria = "ACTIVE"

        if palavra_chave_obrigatoria not in texto:
            messagebox.showerror(
                "ERRO", f"O arquivo ('{nome_arquivo_pdf}') na pasta ('{pasta_selecionada}') não é um PDF válido. Ignorando.")
            continue


except FileNotFoundError as e:
    print(f"Erro: {e}")
    messagebox.showerror("ERRO", f"{e} ENCERRANDO O PROGRAMA.")
    sys.exit()
except Exception as e:
    print(f"Erro inesperado: {e}")
    messagebox.showerror(
        "ERRO", f"Erro inesperado: {e} ENCERRANDO O PROGRAMA.")
    sys.exit()

data_atual = datetime.now()
data_pdf = data_atual.strftime('%d.%m.%Y')
data_formatada = data_atual.strftime('%d.%m.%Y - %H:%M:%S')
nome_arquivo_excel = f'output--{data_pdf}.xlsx'

caminho_completo_excel = os.path.join(os.getcwd(), nome_arquivo_excel)

nomes_arquivos_pdf = [f for f in os.listdir(
    pasta_selecionada) if f.lower().endswith(".pdf")]

data_atual = data_formatada

for nome_arquivo_pdf in nomes_arquivos_pdf:
    caminho_completo_pdf = os.path.join(pasta_selecionada, nome_arquivo_pdf)
    texto = extrair_texto_pdf(caminho_completo_pdf)
    datadeleitura = data_atual
    clientes = extrair_palavra_apos(texto, ' para:')
    cnpj = extrair_outro_dado(texto, 'CNPJ: ')
    condicaodepagamento = extrair_numeros_apos(
        texto, 'Condição de Pagamento :')
    item = extrair_numeros_especificos_apos(texto, 'Valor Total')
    codigo = extrair_proximos_numeros(texto, 'Valor Total')

    if os.path.exists(caminho_completo_excel):
        planilha_existente = load_workbook(caminho_completo_excel)
        folha_existente = planilha_existente.active
    else:
        planilha_existente = Workbook()
        folha_existente = planilha_existente.active
        folha_existente.append([
            'Data de Leitura', 'Cod.Cliente', 'Cliente', 'CNPJ', 'Tipo de Frete',
            'Condição de Pagamento'
        ])
        for celula in folha_existente[1]:
            celula.font = Font(bold=True)

    proxima_linha = folha_existente.max_row + 1

    folha_existente.append([
        datadeleitura if datadeleitura else f"não encontrado{nome_arquivo_pdf}",
        clientes if clientes else f" não encontrado{nome_arquivo_pdf}",
        cnpj if cnpj else f"não encontrado{nome_arquivo_pdf}",
        (condicaodepagamento +
         " dias") if condicaodepagamento else f"não encontrado{nome_arquivo_pdf}",
        item if item else f"não encontrado{nome_arquivo_pdf}",
        codigo if codigo else f"não encontrado{nome_arquivo_pdf}",
    ])

    planilha_existente.save(caminho_completo_excel)
